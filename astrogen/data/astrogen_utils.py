import jellyfish
import difflib
import pandas as pd
import numpy as np
import operator
import re 
import os
from openpyxl import load_workbook
from pathlib import Path

path = Path(__file__).parent / "../../data/external/nombres.csv"
gender_list = pd.read_csv(path)


# COLORS ::::::::::::::::::::::::::::::::::::::::::::::::

class bcolors:
    """
    Get color palette for pretty printing

    This class simply contains a list of predefined colors to
    be used in the visual analysis of strings and publication
    data.

    """
    # ANSI escape sequences
    # Ver también el paquete colorama
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    TST = '\033[31;1m'
    X = '\033[4;95;1m'


# NAME MATCHING :::::::::::::::::::::::::::::::::::::::

def ds(a, b):                                           
    """
    Get distance between two words.

    This function is used to obtain the distance between two names
    or surnames. Uses different distances in word space, namely,
    Damerau Levenshtein distance, Jaro distance, Levenstein
    distance and SequenceMatcher. The later from the difflib package
    and the other ones from the Jellyfish package.

    Args:
        a (string): one of the strings
        b (string): the other string to compare

    Returns:
        res (array): Numpy array with the list of distances
           between the two words.

    """
    d1 = jellyfish.damerau_levenshtein_distance(a, b)
    d2 = jellyfish.jaro_distance(a, b)     
    d3 = jellyfish.levenshtein_distance(a, b)                          
    s = difflib.SequenceMatcher(None, a, b)                            
    d4 = s.ratio()                                                
    res = np.array([d1, d2, d3, d4])
    return res

def ds1(s1, s2):
    """
    Get distance between two words.

    This function is used to obtain the distance between two names
    or surnames. Uses different distances in word space, namely,
    Damerau Levenshtein distance, Jaro distance, Levenstein
    distance and SequenceMatcher. The later from the difflib package
    and the other ones from the Jellyfish package.

    Args:
        a (string): one of the strings
        b (string): the other string to compare

    Returns:
        res (array): Numpy array with the list of distances
           between the two words.

    """ 
    s1l = s1.lower().split()
    s2l = s2.lower().split()
    n1 = len(s1l)
    n2 = len(s2l)
    if n1 > 1: s1l.append(s1.lower())
    if n2 > 1: s2l.append(s2.lower())
    dm = 99
    for p1 in s1l:
        for p2 in s2l:
            s = difflib.SequenceMatcher(None, p1, p2)
            d = 1 - s.ratio()                                                
            dm = min(dm, d)
    return dm

def ds2(ap1, ap2, nom1, nom2):
    """
    Get distance between two words.

    This function is used to obtain the distance between two names
    or surnames. Uses different distances in word space, namely,
    Damerau Levenshtein distance, Jaro distance, Levenstein
    distance and SequenceMatcher. The later from the difflib package
    and the other ones from the Jellyfish package.

    Args:
        a (string): one of the strings
        b (string): the other string to compare

    Returns:
        res (array): Numpy array with the list of distances
           between the two words.

    """    
    d_apel = ds1(ap1, ap2)
    d_nomb = ds1(nom1, nom2)
    names_dist = np.sqrt(d_apel**2 + d_nomb**2)
    return names_dist



def initials(initials, string):
    """
    Check if the initials of two names coincide.

    e.g.:

    initials = 'Juan Carlos'; string='Juan' --> True

    initials = 'Juan Carlos'; string='Juan José' --> False

    initials = 'Juan Carlos'; string='Jacinto' --> True

    Args:
        initials (string): source string for the initials
        string (string): full names

    Returns:
        boo (bool): whether the initials are accepted

    Notes:

    The criteria for the string matching is the following:

    """
    Li = [x[0] for x in initials.lower().replace('.', ' ').split()]
    Ln = [x[0] for x in string.lower().replace('.', ' ').split()]
    ni = len(Li)
    nn = len(Ln)
    boo = Li==Ln
    if (ni==1 or nn==1) and ni!=nn:
        boo = Li[0]==Ln[0]
    return boo

def getinitials(nombre):
    """
    Get the initials of a full name

    e.g.: 'Jose Facundo' --> 'J. F.'

    Args:

    Returns:

    """
    res = ' '.join([a[0].upper()+'.' for a in nombre.split()])
    return res
 
def pickone(df, au, sift):
    """
    de una lista de autores en un dataframe "df" 
    elige el que está más cerca de un autor "au"
    y devuelve un array booleano que es todo falso
    savo uno (el autor elegido).
    """
    a1, n1 = au.split(',')
    dopt = 99
    ind = 0
    for i, r in enumerate(df.iterrows()):
        if not sift[i]: continue
        a2, n2 = r[1][0], r[1][1]
        d = ds2(a1, a2, n1, n2)
        if d < dopt:
            dopt = d
            ind = i
    sift = [False]*len(sift)
    sift[ind] = True
    return sift
 

# TEXT MANIPULATION ::::::::::::::::::::::::::::::::::::::::::::

def clean_text(txt):
    txt = re.sub("[^a-záéíóúñüäë]", " ", txt.lower())
    txt = re.sub(' +',' ', txt)
    return txt.strip().split()

def df_to_dict(df, key_column, val_column):
    """convierte dos pandas series en un diccionario"""
    xkey = df[key_column].tolist()
    xval = df[val_column].tolist()
    return dict(zip(xkey,xval))

gender_list = df_to_dict(gender_list, key_column='nombre', val_column='genero')

# GENDER DETECTION :::::::::::::::::::::::::::::::::::::::::::::

def get_gender2(names):
    names = clean_text(names)
    names = [x for x in names if gender_list.get(x,'a') != 'a']
    gender ={'m':0, 'f':0, 'a':0}
    for i, name in enumerate(names):
        g = gender_list.get(name,'a')
        gender[g] += 1
        gender[g] += 2 if len(names) > 1 and i == 0 and g != 'a' else 0 
    gender['a'] = 0 if (gender['f']+gender['m']) > 0 else 1
    return max(gender.items(), key=operator.itemgetter(1))[0]

 
 
# XLSX WRITERS :::::::::::::::::::::::::::::::::::::::::::::::::

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
    (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
    (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
    Per default (startrow=None) calculate the last row
    in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
    before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
    [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_df_to_excel('d:/temp/test.xlsx', df)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2', 
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name, 
            startrow=startrow if startrow is not None else 0, 
            **to_excel_kwargs)
        return
    
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)
    
    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)
    
    # copy existing sheets
    writer.sheets = {ws.title:ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
